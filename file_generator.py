import os
import re
from openpyxl import load_workbook
import pandas as pd
from pathlib import PurePath
from transliterate import translit
from image_generator import image_function


def file_function(parameters):
    path_to_table = parameters.name
    translations_table = load_workbook(filename=path_to_table, data_only=True)
    current_sheet = parameters.sheet
    if current_sheet.isdigit():
        translations_sheet = translations_table[translations_table.sheetnames[int(current_sheet)]]
    else:
        translations_sheet = translations_table[current_sheet]

    titles = pd.read_excel(path_to_table)
    number_of_rows = len(titles) + 1
    number_of_columns = len(titles.columns)
    folder_path = parameters.path

    for i in range(2, number_of_columns + 1):
        folder = PurePath(folder_path, str(translations_sheet.cell(row=1, column=i).value))
        os.makedirs(folder, exist_ok=True)

    for i in range(2, number_of_rows + 1):
        for j in range(2, len(titles.columns) + 1):
            folder = PurePath(folder_path, str(translations_sheet.cell(row=1, column=j).value))
            name = translit(translations_sheet.cell(row=i, column=1).value, language_code='ru',
                            reversed=True)
            name = re.sub(r'[\\/*?+:"<>|]', '', name)
            image_function(translations_sheet.cell(row=i, column=j).value,
                           str(PurePath(folder, name + '.png')), parameters.back, parameters.text, parameters.font,
                           parameters.size, parameters.wpad, parameters.hpad)
