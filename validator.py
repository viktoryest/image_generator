import os.path
from openpyxl import load_workbook


def is_valid(parameters):
    name = parameters.name
    sheet = parameters.sheet
    font = parameters.font
    back = parameters.back
    text = parameters.text
    size = parameters.size
    wpad = parameters.wpad
    hpad = parameters.hpad
    table_extension = name[-4:]
    table_extension_list = ['xlsx', 'xlsx', 'xlsm', 'xltx', 'xltm']
    if table_extension not in table_extension_list:
        return 'Invalid extension for the table. Allowed formats: .xlsx, .xlsx, .xlsm, .xltx, .xltm'
    if not os.path.isfile(name):
        return "Table file doesn't exist. Please, check the path"

    table = load_workbook(filename=name, data_only=True)
    number_of_sheets = len(table.sheetnames)
    if str(sheet).isdigit():
        if int(sheet) >= number_of_sheets:
            return 'Invalid sheet value. Note that numbering starts at zero'
    elif not str(sheet).isdigit():
        if sheet not in table.sheetnames:
            return 'The specified sheet does not exist'

    font_extension_list = ['.BDF', '.CHR', '.DFONT', '.EOT', '.ETX', '.FNT', '.FON', '.FOT', '.GDR', '.GF',
                           '.GXF', '.MCF', '.MF', '.OTF', '.PF', '.PFA', '.PFB', '.PFM', '.PFR', '.SFD',
                           '.SUIT', '.TFM', '.TTC', '.TTF', '.VLW', '.WOFF', '.XFN', '.XFT']
    font_extension_str = ', '.join(font_extension_list)
    for i in range(len(font)):
        if '.' not in font:
            return f'Invalid font name. Allowed formats: {font_extension_str}'
        if font[i] == '.':
            point = i
            font_extension = font[point:]
            if font_extension.upper() not in font_extension_list:
                return f"Invalid font name. Allowed formats: {font_extension_str}"

    if back[0] != '#':
        if back[:4] != 'rgb' or len(back) > 20:
            return '''Invalid background colour. In RGB colour must have format "rgb'(x, y, z)'"'''
    elif back[0] == '#':
        if len(back) > 7:
            return 'Invalid background colour. In hex colour must have format "#xxxxxx" or "#xxx"'

    if text[0] != '#':
        if text[:4] != 'rgb' or len(text) > 20:
            return '''Invalid text colour. In RGB colour must have format "rgb'(x, y, z)'"'''
    elif back[0] == '#':
        if len(back) > 7:
            return 'Invalid background colour. In hex colour must have format "#xxxxxx" or "#xxx"'

    if type(size) is not int or size <= 0 or type(size) is not int and size <= 0:
        return 'Invalid text size (must be a natural number)'

    if type(wpad) is not int or wpad <= 0 or (type(wpad) is not int and wpad <= 0):
        return 'Invalid width padding (must be a natural number)'

    if type(hpad) is not int or hpad <= 0 or (type(hpad) is not int and hpad <= 0):
        return 'Invalid height padding (must be a natural number)'
    else:
        return True
